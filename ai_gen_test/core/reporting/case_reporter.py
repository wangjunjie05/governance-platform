import json
import os
import re
from collections import defaultdict
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=False)
    except Exception:
        return str(value)


def _bool_text(value: Any) -> str:
    if isinstance(value, bool):
        return "是" if value else "否"
    return value




def _seconds_text(value: Any) -> Any:
    """Excel 展示用耗时。

    特殊下载/导出接口可能走稳定生成逻辑，耗时小于 0.01 秒时，
    原来四舍五入会显示 0，容易误解为没有生成。这里仅调整展示，
    不改变内部统计用的 generate_time_s 数值。
    """
    if value in (None, ""):
        return ""
    try:
        seconds = float(value)
    except Exception:
        return value
    if 0 <= seconds < 0.01:
        return "<0.01"
    return round(seconds, 2)

def _result_text(value: str) -> str:
    if value == "passed":
        return "通过"
    if value == "failed":
        return "失败"
    if value in {"error", "errors"}:
        return "错误"
    if value == "skipped":
        return "跳过"
    if value:
        return value
    return ""




_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
_MAX_CELL_TEXT_LEN = 32000


def _clean_cell_value(value: Any) -> Any:
    """清理写入 Excel 的单元格内容。

    openpyxl 不允许控制字符写入单元格。代理日志里如果误采集了
    Excel/ZIP/二进制响应体，直接写入会触发 IllegalCharacterError。
    这里统一转成安全字符串，并限制长度。
    """
    value = _bool_text(value)

    if value is None:
        return ""

    if isinstance(value, (int, float)):
        return value

    if isinstance(value, bool):
        return "是" if value else "否"

    if isinstance(value, bytes):
        value = value.decode("utf-8", errors="replace")

    text = str(value)
    text = _ILLEGAL_XLSX_RE.sub("", text)

    if len(text) > _MAX_CELL_TEXT_LEN:
        text = text[:_MAX_CELL_TEXT_LEN] + "\n...（内容过长，已截断）"

    return text


def load_case_rows(jsonl_path: str) -> List[Dict[str, Any]]:
    if not jsonl_path or not os.path.exists(jsonl_path):
        return []
    rows: List[Dict[str, Any]] = []
    with open(jsonl_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except Exception:
                continue
    return rows



def _case_signature(row: Dict[str, Any]) -> str:
    """用于报告统计去重：同一接口、同一请求、同一响应、同一用例类型视为同一条用例。"""
    def stable(value: Any) -> str:
        try:
            return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
        except Exception:
            return str(value)
    return "|".join([
        str(row.get("file_name", "")),
        str(row.get("case_type", "")),
        str(row.get("method", "")),
        str(row.get("url", "")),
        stable(row.get("params")),
        stable(row.get("request_body")),
        str(row.get("response_status", "")),
        stable(row.get("response_body")),
    ])


def dedupe_case_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """去除 AI 用例和自动补充用例产生的重复记录，避免统计重复。"""
    result: List[Dict[str, Any]] = []
    seen = set()
    for row in rows or []:
        sig = _case_signature(row)
        if sig in seen:
            continue
        seen.add(sig)
        result.append(row)
    return result


def summarize_case_rows(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    rows = dedupe_case_rows(rows)
    summary = {
        "normal_case_count": 0,
        "exception_case_count": 0,
        "passed": 0,
        "failed": 0,
        "errors": 0,
        "skipped": 0,
    }
    for row in rows:
        if row.get("case_type") == "异常":
            summary["exception_case_count"] += 1
        else:
            summary["normal_case_count"] += 1

        result = row.get("result")
        if result == "passed":
            summary["passed"] += 1
        elif result == "failed":
            summary["failed"] += 1
        elif result == "skipped":
            summary["skipped"] += 1
        elif result in {"error", "errors"}:
            summary["errors"] += 1
        else:
            summary["errors"] += 1
    return summary


def _append_sheet_rows(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append([_clean_cell_value(v) for v in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_case_summary_rows(all_case_rows: List[Dict[str, Any]]) -> List[List[Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in all_case_rows:
        grouped[row.get("file_name", "unknown")].append(row)

    rows: List[List[Any]] = []
    for file_name, case_rows in sorted(grouped.items()):
        case_rows = dedupe_case_rows(case_rows)
        summary = summarize_case_rows(case_rows)
        total = len(case_rows)
        passed = summary["passed"]
        rate = round(passed / total * 100, 2) if total else 0
        rows.append([
            file_name,
            total,
            summary["normal_case_count"],
            summary["exception_case_count"],
            summary["passed"],
            summary["failed"],
            summary["errors"],
            summary["skipped"],
            f"{rate}%",
        ])
    return rows


def build_case_detail_rows(all_case_rows: List[Dict[str, Any]]) -> List[List[Any]]:
    all_case_rows = dedupe_case_rows(all_case_rows)
    rows: List[List[Any]] = []
    for row in all_case_rows:
        rows.append([
            row.get("file_name", ""),
            row.get("test_name", ""),
            row.get("case_type", ""),
            _result_text(row.get("result", "")),
            row.get("method", ""),
            row.get("url", ""),
            _json_text(row.get("params")),
            _json_text(row.get("request_body")),
            row.get("response_status", ""),
            _json_text(row.get("response_body")),
            row.get("duration_s", ""),
            row.get("error_msg", ""),
        ])
    return rows


def build_interface_result_rows(result_rows: List[Dict[str, Any]]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for row in result_rows:
        rows.append([
            row.get("method", ""),
            row.get("path", ""),
            row.get("file_name", ""),
            row.get("model", ""),
            _seconds_text(row.get("generate_time_s", "")),
            row.get("generate_success", ""),
            row.get("syntax_success", ""),
            row.get("pytest_success", ""),
            row.get("fix_retry_used", ""),
            row.get("fallback_used", ""),
            row.get("fallback_reason", ""),
            row.get("code_length", ""),
            row.get("data_source", ""),
            row.get("knowledge_base", ""),
            row.get("actual_knowledge_base", ""),
            row.get("generation_mode", ""),
            row.get("exception_mode", ""),
            row.get("log_case_count", ""),
            row.get("code_scan_matched", ""),
            row.get("normal_case_count", 0),
            row.get("exception_case_count", 0),
            row.get("exception_case_rate", ""),
            row.get("pytest_passed_count", 0),
            row.get("pytest_failed_count", 0),
            row.get("pytest_error_count", 0),
            row.get("pytest_skipped_count", 0),
            row.get("script_executable_rate", ""),
            row.get("pytest_case_pass_rate", ""),
            row.get("pytest_report_file", ""),
            row.get("error_type", ""),
            row.get("error_msg", ""),
        ])
    return rows


def write_run_report_xlsx(result_rows: List[Dict[str, Any]], all_case_rows: List[Dict[str, Any]], xlsx_path: str) -> None:
    """写入本轮运行报告。

    Excel 支持多个 sheet，比 CSV 分段更适合人工查看：
    1. 接口生成结果
    2. 接口用例汇总
    3. 用例执行明细
    """
    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb = Workbook()

    ws_interface = wb.active
    ws_interface.title = "接口生成结果"
    interface_headers = [
        "请求方法", "接口路径", "测试脚本文件", "模型", "生成耗时秒", "是否生成成功",
        "语法是否通过", "pytest接口是否通过", "是否触发修复", "是否触发兜底", "兜底原因", "代码长度", "数据来源",
        "期望知识库", "实际命中来源", "生成模式", "异常用例模式", "日志用例数", "代码扫描是否命中",
        "正常用例数", "异常用例数", "异常用例占比", "pytest用例通过数", "pytest用例失败数",
        "pytest用例错误数", "pytest用例跳过数", "脚本可执行率", "用例通过率", "pytest文本报告路径", "错误类型", "错误信息",
    ]
    _append_sheet_rows(ws_interface, interface_headers, build_interface_result_rows(result_rows))

    ws_summary = wb.create_sheet("接口用例汇总")
    case_summary_headers = ["接口文件", "用例总数", "正常用例数", "异常用例数", "通过数", "失败数", "错误数", "跳过数", "通过率"]
    _append_sheet_rows(ws_summary, case_summary_headers, build_case_summary_rows(all_case_rows))

    ws_detail = wb.create_sheet("用例执行明细")
    case_detail_headers = [
        "接口文件", "用例名称", "用例类型", "执行结果", "请求方法", "请求URL",
        "请求参数", "请求Body", "响应状态码", "响应Body", "耗时秒", "失败原因",
    ]
    _append_sheet_rows(ws_detail, case_detail_headers, build_case_detail_rows(all_case_rows))

    wb.save(xlsx_path)
